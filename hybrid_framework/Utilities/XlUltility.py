import openpyxl

# get row
def GetRow(xlpath, sheet):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.max_row

# get col
def GetCol(xlpath, sheet):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.max_column

# read data
def ReadData(xlpath, sheet, r, c):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.cell(r,c).value

# write data
def WriteData(xlpath, sheet, r, c, data):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    sh.cell(r,c).value=data
    wb.save(xlpath)
