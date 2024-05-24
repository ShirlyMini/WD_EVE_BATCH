import openpyxl

## open xlsx or xlx(excel/workbook)--> sheet-->row and col

### write(save xl is mandatory)
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# # wb.create_sheet("sheet1", 1)
# sh = wb['Sheet']
# sh.cell(1,1).value="SNO"
# sh.cell(1,2).value="name"
# sh.cell(1,3).value="mark"
# sh.cell(2,1).value="1"
# sh.cell(2,2).value="john"
# sh.cell(2,3).value="80"
# sh.cell(3,1).value="2"
# sh.cell(3,2).value="scott"
# sh.cell(3,3).value="90"
# wb.save("demo_xl.xlsx")


### read(no need to save xl)
wb = openpyxl.load_workbook(r".\demo_xl.xlsx")
sh = wb['Sheet']
print(sh.max_row)
print(sh.max_column)
# print(sh.cell(1,1).value)
for r in range(1, sh.max_row+1):#3-0 to 2
    for c in range(1,sh.max_column+1):
        print(sh.cell(r,c).value)

